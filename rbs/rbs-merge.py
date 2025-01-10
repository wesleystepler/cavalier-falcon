from difflib import SequenceMatcher
from openpyxl import load_workbook
import time

workbook = load_workbook(filename='rbs-report.xlsx', data_only=True)
sheet = workbook.active
for row in range(12, workbook.max_row):
    for col in workbook.iter_cols(1, workbook.max_column):
        print(col[row].value)

# The list that will store detected possible duplicates
possible_duplicates = ["Test1", "Test2", "Test3", "Test4"]
merged_records = []
while len(possible_duplicates) >= 2:
    print("Possible duplicates detected:")
    print()
    print(f"Record #1: {possible_duplicates[0]}")
    print(f"Record #2: {possible_duplicates[1]}")
    print()
    # This flag will be false until the user submits a valid input for merging
    valid_response = False
    while not valid_response:
        merge = input("Merge these records? (Y/N)  ")
        if merge == "Y" or merge == "y":
            valid_response = True
            # This flag will be false until the user submits a valid input for record selection
            valid_record = False
            while not valid_record:
                parent_record = input(f"Which record name would you like to keep? 1: {possible_duplicates[0]}, or 2: {possible_duplicates[1]}? Please response with '1' or '2'.  ")
                if parent_record == "1":
                    # Append the chosen record name to merged_records, and delete each of the accounts from the duplicates list
                    merged_records.append(possible_duplicates[0])
                    print(f"Merged records {possible_duplicates[0]} and {possible_duplicates[1]} into {possible_duplicates[0]}")
                    time.sleep(0.5)
                    possible_duplicates.remove(possible_duplicates[0])
                    possible_duplicates.remove(possible_duplicates[0])
                    valid_record = True
                    
                elif parent_record == "2":
                    merged_records.append(possible_duplicates[1])
                    print(f"Merged records {possible_duplicates[0]} and {possible_duplicates[1]} into {possible_duplicates[1]}")
                    time.sleep(0.5)
                    possible_duplicates.remove(possible_duplicates[0])
                    possible_duplicates.remove(possible_duplicates[0])
                    valid_record = True

                else:
                    print("Invalid Input. Please response with either '1' or '2'")
                    time.sleep(0.5)

        elif merge == "N" or merge == "n":
            valid_response = True
            possible_duplicates.remove([0])
            possible_duplicates.remove([0])

        else:
            print("Invalid Input. Please response with either 'Y' or 'N'")

print(merged_records)
