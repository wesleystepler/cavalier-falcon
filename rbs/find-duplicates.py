from openpyxl import load_workbook
from difflib import SequenceMatcher

workbook = load_workbook(filename='rbs-report.xlsx', data_only=True)
sheet = workbook.active

def compare_strings(a: str, b: str):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def compare_words(a, b):
    a_split = a.split()
    b_split = b.split()

def find_duplicates(sheet):
    row, col = 12, 5
    potential_duplicates = []
    visited_cells = []
    while sheet.cell(row, col).value != None:
        c1 = sheet.cell(row, col)
        c1_val = sheet.cell(row, col).value
        i, j, = row+1, col
        while sheet.cell(i, j).value != None:
            print(i, j)
            similar = False
            c2 = sheet.cell(i, j)
            c2_val = sheet.cell(i, j).value
            if c1_val == c2_val and c1 != c2:
                similar = True
            
            elif compare_strings(c1_val, c2_val) > 0.85:
                similar = True

            if similar:
                if c1 not in visited_cells:
                    potential_duplicates.append(c1_val)
                    visited_cells.append(c1)
                if c2 not in visited_cells:
                    potential_duplicates.append(c2_val)
                    visited_cells.append(c2)
            i += 1
        row += 1

    return potential_duplicates

print(find_duplicates(sheet))


